from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from constant import globalconstant
import pytest
import openpyxl
import json



class Test_work():

    def setup_method(self):
        self.driver=webdriver.Chrome()
        self.driver.get(globalconstant.url)
        self.driver.maximize_window()
        sleep(5)
    def teardown_method(self):
        self.driver.quit()



    def getData():
        excelFile = openpyxl.load_workbook(globalconstant.excelUzantisi) 
        sheet = excelFile["Sheet1"] 
        rows = sheet.max_row 
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value 
            password = sheet.cell(i,2).value
            data.append((username,password))
    
        return data
    
    @pytest.mark.parametrize("username,password",getData())
    def test_invalid_login(self,username,password):
        
        user = self.driver.find_element(By.ID,globalconstant.username_id)
        user.send_keys(username)
        passw = self.driver.find_element(By.ID,globalconstant.password_İD)
        passw.send_keys(password)
        loginButton = self.driver.find_element(By.ID,globalconstant.logginButton_id)
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH,globalconstant.errmassage_xpath)
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"

    

    def test_incorrectConditions(self):
        username=self.driver.find_element(By.ID,globalconstant.username_id)
        username.send_keys("")
        password=self.driver.find_element(By.ID,globalconstant.password_İD)
        password.send_keys("")
        logginButton=self.driver.find_element(By.ID,globalconstant.logginButton_id)
        logginButton.click()
        sleep(2)
        errorMassage=self.driver.find_element(By.XPATH,globalconstant.errmassage_empty_xpath)
        assert errorMassage.text=="Epic sadface: Username is required"


    def getData2():
        excelFile = openpyxl.load_workbook(globalconstant.excelLogin) 
        sheet = excelFile["Sayfa1"] 
        rows = sheet.max_row 
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value 
            password = sheet.cell(i,2).value
            data.append((username,password))
    
        return data
        
    @pytest.mark.parametrize("username,password",getData2())

    def test_login(self,username,password):
        userr=self.driver.find_element(By.ID,globalconstant.username_id)
        userr.send_keys(username)
        sleep(2)
        passwo=self.driver.find_element(By.ID,globalconstant.password_İD)
        passwo.send_keys(password)
        sleep(3)
        logginButton=self.driver.find_element(By.ID,globalconstant.logginButton_id)
        logginButton.click()
        sleep(3)
        baslik=self.driver.find_element(By.XPATH,globalconstant.title_path)
        assert baslik.text=="Swag Labs"


    def veritabanini_yukle():
     with open("veritabani/users.json") as dosya:
        data = json.load(dosya)
    # "account" anahtarının altındaki kullanıcı bilgilerini al
        accounts = data.get("account", [])
    # Kullanıcı adı ve şifreleri çıkar
        usernames = [account["username"] for account in accounts]
        passwords = [account["password"] for account in accounts]
    # Kullanıcı adı ve şifreleri birleştirip zip ile tuple oluştur
        return zip(usernames, passwords)



    @pytest.mark.parametrize("username, password", veritabanini_yukle())
    def test_incorrectUsername(self, username, password):
        user = self.driver.find_element(By.ID, globalconstant.username_id)
        user.send_keys(username)
        sleep(2)
        passw = self.driver.find_element(By.ID, globalconstant.password_İD)
        passw.send_keys(password)
        logginButton = self.driver.find_element(By.ID, globalconstant.logginButton_id)
        logginButton.click()
        sleep(3)
        errorMassage = self.driver.find_element(By.XPATH, globalconstant.errorLocked)
        assert errorMassage.text == "Epic sadface: Sorry, this user has been locked out."


    
    



   
        

       






        